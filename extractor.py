import argparse
import json
import os
from pathlib import Path
from typing import Iterable

import fitz  # PyMuPDF
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from utils import parse_pdf_date, get_annot_color, extract_highlighted_text

ANNOT_TYPE_MAP = {
    "Text": "Sticky Note",
    "Highlight": "Highlight",
    "Underline": "Underline",
    "StrikeOut": "Strikethrough",
    "Squiggly": "Squiggly",
    "FreeText": "FreeText",
    "Ink": "Ink",
    "Line": "Line",
    "Square": "Square",
    "Circle": "Circle",
    "Polygon": "Polygon",
    "FileAttachment": "File Attachment",
    "Stamp": "Stamp",
    "Redact": "Redaction",
}

TYPE_FILL_MAP = {
    "Highlight": "FFF59D",
    "Sticky Note": "90CAF9",
    "FreeText": "A5D6A7",
    "Strikethrough": "EF9A9A",
    "Redaction": "EF9A9A",
}

DEFAULT_COLUMNS = [
    "source_file",
    "page_number",
    "annotation_type",
    "author",
    "date_created",
    "date_modified",
    "subject",
    "contents",
    "highlighted_text",
    "color",
    "rect",
    "flags",
    "popup_open",
]


def collect_pdf_paths(files: Iterable[str], folder: str | None) -> list[Path]:
    paths = [Path(p) for p in files] if files else []
    if folder:
        folder_path = Path(folder)
        if folder_path.is_dir():
            paths.extend(sorted(folder_path.glob("*.pdf")))
    unique = []
    seen = set()
    for path in paths:
        if path.exists() and path.suffix.lower() == ".pdf" and path not in seen:
            unique.append(path)
            seen.add(path)
    return unique


def extract_annotations(
    pdf_path: Path,
    include_color: bool,
    include_text: bool,
    skip_empty: bool,
) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    warnings: list[str] = []

    try:
        doc = fitz.open(pdf_path)
    except Exception as exc:
        warnings.append(f"Failed to open {pdf_path.name}: {exc}")
        return rows, warnings

    if doc.is_encrypted and not doc.authenticate(""):
        warnings.append(f"Skipped encrypted file: {pdf_path.name}")
        doc.close()
        return rows, warnings

    for page_num, page in enumerate(doc, start=1):
        try:
            annots = list(page.annots() or [])
        except Exception as exc:
            warnings.append(f"Skipped page {page_num} in {pdf_path.name}: {exc}")
            continue

        for annot in annots:
            try:
                info = annot.info or {}
                raw_type = annot.type[1] if annot.type else "Unknown"
                annotation_type = ANNOT_TYPE_MAP.get(raw_type, raw_type)
                highlighted_text = ""
                if include_text and raw_type in {"Highlight", "Underline", "StrikeOut", "Squiggly"}:
                    highlighted_text = extract_highlighted_text(page, annot)
                color = get_annot_color(annot) if include_color else ""

                rect_value = json.dumps(list(annot.rect)) if annot.rect else ""
                row = {
                    "source_file": pdf_path.name,
                    "page_number": page_num,
                    "annotation_type": annotation_type,
                    "author": info.get("title", ""),
                    "date_created": parse_pdf_date(info.get("creationDate", "")),
                    "date_modified": parse_pdf_date(info.get("modDate", "")),
                    "subject": info.get("subject", ""),
                    "contents": info.get("content", ""),
                    "highlighted_text": highlighted_text,
                    "color": color,
                    "rect": rect_value,
                    "flags": annot.flags,
                    "popup_open": annot.is_open,
                }
                rows.append(row)
            except Exception as exc:
                warnings.append(
                    f"Skipped annotation on page {page_num} in {pdf_path.name}: {exc}"
                )

    doc.close()

    if not rows and not skip_empty:
        rows.append(
            {
                "source_file": pdf_path.name,
                "page_number": "",
                "annotation_type": "None",
                "author": "",
                "date_created": "",
                "date_modified": "",
                "subject": "",
                "contents": "No annotations found",
                "highlighted_text": "",
                "color": "",
                "rect": "",
                "flags": "",
                "popup_open": "",
            }
        )

    return rows, warnings


def write_csv(output_path: Path, df: pd.DataFrame, overwrite: bool) -> None:
    exists = output_path.exists()
    mode = "w" if overwrite or not exists else "a"
    header = overwrite or not exists
    df.to_csv(output_path, mode=mode, index=False, header=header)


def apply_xlsx_formatting(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E0E0E0")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.freeze_panes = "A2"

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="FAFAFA")
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    try:
        type_col_index = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0].index(
            "annotation_type"
        ) + 1
    except ValueError:
        return

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=type_col_index)
        fill_color = TYPE_FILL_MAP.get(cell.value)
        if fill_color:
            cell.fill = PatternFill("solid", fgColor=fill_color)

    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def write_xlsx(
    output_path: Path,
    df: pd.DataFrame,
    overwrite: bool,
    sheet_per_file: bool,
) -> None:
    if output_path.exists() and not overwrite:
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_per_file:
        for source_file, group in df.groupby("source_file"):
            sheet_name = source_file[:31]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(DEFAULT_COLUMNS)
            for row in group[DEFAULT_COLUMNS].itertuples(index=False):
                ws.append(list(row))
            apply_xlsx_formatting(ws)
    else:
        sheet_name = "Annotations"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(DEFAULT_COLUMNS)
        for row in df[DEFAULT_COLUMNS].itertuples(index=False):
            ws.append(list(row))
        apply_xlsx_formatting(ws)

    wb.save(output_path)


def write_json(output_path: Path, rows: list[dict], overwrite: bool) -> None:
    data = rows
    if output_path.exists() and not overwrite:
        try:
            existing = json.loads(output_path.read_text(encoding="utf-8"))
            if isinstance(existing, list):
                data = existing + rows
        except Exception:
            pass
    output_path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract PDF annotations and export to CSV/XLSX/JSON."
    )
    parser.add_argument("files", nargs="*", help="PDF file paths")
    parser.add_argument("--folder", help="Folder containing PDF files")
    parser.add_argument(
        "--output",
        default="annotations_output.xlsx",
        help="Output file path (default: annotations_output.xlsx)",
    )
    parser.add_argument("--format", choices=["csv", "xlsx", "json"], help="Force output format")
    parser.add_argument("--no-color", action="store_true", help="Disable color extraction")
    parser.add_argument("--no-text", action="store_true", help="Skip highlighted text extraction")
    parser.add_argument(
        "--sheet-per-file",
        action="store_true",
        help="(XLSX only) Put each PDF on its own sheet",
    )
    parser.add_argument("--verbose", action="store_true", help="Verbose output")
    parser.add_argument(
        "--skip-empty",
        action="store_true",
        help="Skip PDFs with no annotations",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite output file instead of appending",
    )
    return parser.parse_args()


def determine_format(output: str, forced: str | None) -> str:
    if forced:
        return forced
    ext = Path(output).suffix.lower()
    if ext == ".csv":
        return "csv"
    if ext == ".json":
        return "json"
    return "xlsx"


def main() -> None:
    args = parse_args()
    pdf_paths = collect_pdf_paths(args.files, args.folder)
    if not pdf_paths:
        print("No PDF files found.")
        return

    output_path = Path(args.output)
    output_format = determine_format(args.output, args.format)

    all_rows: list[dict] = []
    warnings: list[str] = []

    use_progress = len(pdf_paths) > 1
    iterator = tqdm(pdf_paths, desc="Processing PDFs") if use_progress else pdf_paths

    for pdf_path in iterator:
        rows, file_warnings = extract_annotations(
            pdf_path,
            include_color=not args.no_color,
            include_text=not args.no_text,
            skip_empty=args.skip_empty,
        )
        all_rows.extend(rows)
        warnings.extend(file_warnings)
        if args.verbose:
            print(f"{pdf_path.name}: {len(rows)} annotations")

    if not all_rows:
        print("No annotations extracted.")
        if warnings:
            for warning in warnings:
                print(f"Warning: {warning}")
        return

    df = pd.DataFrame(all_rows)
    df = df[DEFAULT_COLUMNS]
    df.sort_values(by=["source_file", "page_number", "date_created"], inplace=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_format == "csv":
        write_csv(output_path, df, args.overwrite)
    elif output_format == "json":
        write_json(output_path, all_rows, args.overwrite)
    else:
        write_xlsx(output_path, df, args.overwrite, args.sheet_per_file)

    print(
        f"Done. Extracted {len(df)} annotations from {len(pdf_paths)} PDFs -> {output_path}"
    )
    for warning in warnings:
        print(f"Warning: {warning}")


if __name__ == "__main__":
    main()
